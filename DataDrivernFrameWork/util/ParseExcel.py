# encoding = utf-8
import openpyxl
from openpyxl.styles import Border, Side, Font
import time

class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None) # 设置字体颜色
        # 颜色对应RGB值
        self.RGBDict = {"red": "FFFF3030", "Green": "FF008B00"}

    def loadWorkBook(self, excelPathAndName):
        # 将Excel文件加载到内存，并获取其workbook
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except:
            print("Excel文件加载到内存抛异常了")
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self, sheetName):
        #根据sheet名称获取该sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except:
            print("根据sheet名称获取该sheet对象时抛异常了")

    def getSheetByIndex(self, sheetIndex):
        # 根据sheet索引号获取该sheet对象
        try:
            sheetName = self.workbook.get_sheet_names()[sheetIndex]

        except:
            print("根据sheet索引号获取该sheet对象时抛异常了")
        sheet = self.workbook.get_sheet_by_name(sheetName)
        return sheet

    def getRowsNumber(self, sheet):
        # 获取sheet 中有数据区域的结束号
        return sheet.max_row

    def getColsNumber(self, sheet):
        # 获取sheet 中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        # 获取sheet 中有数据区域的开始行号
        return sheet.min_row

    def getStartColNumber(self, sheet):
        # 获取sheet 中有数据区域的开始列号
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        # 获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple
        # 下标从1开始，sheet.rows[1]表示第一行
        try:
            return sheet.rows[rowNo-1]
        except:
            print("获取sheet中某一行抛异常了")

    def getColumn(self, sheet, colNo):
        # 获取sheet中某一列，返回的是这一列所有的数据内容组成的tuple
        # 下标从1开始，sheet.rows[1]表示第一列
        try:
            return sheet.columns[colNo - 1]
        except:
            print("获取sheet中某一列抛异常了")

    def getCellOfValue(self, sheet, coordinate=None, rowNo = None, colsNo = None):
        # 根据单元格所在的位置索引获取该单元格中的值，下标从1开始，
        # sheet.cell(row =1 ,column = 1).value,表示Excel中第一行第一列的值
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except :
                print("getCellOfValue方法抛异常啦")
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo).value
            except :
                print("getCellOfValue方法抛异常啦")
        else:
            raise Exception("Insufficient Coordinates of cell")

    def getCellOfObject(self, sheet, coordinate=None, rowNo = None, colsNo = None):
        # 获取某个单元格的的对象，可以根据单元格所在位置的数字索引，
        # 也可以直接根据Excel 中的单元格的编码及坐标
        # 如 getCellObject(sheet, coordinate = "Al")or getCellObject(sheet, rowNo=1, colsNo=2)
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate)
            except :
                raise Exception
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo)
            except:
                raise Exception
        else:
            raise Exception("Insufficient Coordinates of cell")

    def writeCell(self, sheet, content, coordinate=None, rowNo = None, colsNo = None, style=None):
        # 根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据，
        # 下标从1开始，参数style表示字体的颜色的名字，比如red，green
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except:
                raise Exception
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                if style:
                    sheet.cell(row = rowNo, column = colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except:
                raise Exception
        else:
            raise Exception("Insufficient Coordinates of cell")
    def writeCellCurrentTime(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        # 写入当前的时间，鞋标从1开始
        now = int(time.time()) # 显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("% Y - % m - % d % H: % M :% S",timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except :
                raise Exception
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value =currentTime
                self.workbook.save(self.excelFile)
            except :
                raise Exception
        else:
            raise Exception("Insufficient Coordinates of cell")
if __name__ == "__main__":
    pe = ParseExcel()
    # 先创建excel文件 C:\devTools\repository\test\DataDrivernFrameWork\testData
    pe.loadWorkBook(u"C:\\devTools\\repository\\test\\DataDrivernFrameWork\\testData\\163邮箱联系人.xlsx")
    print("通过名称获取sheet对象的名字：", pe.getSheetByName(u"联系人").title)
    print("通过index序号获取sheet对象的名字：", pe.getSheetByIndex(0).title)
    sheet = pe.getSheetByIndex(0)
    print(type(sheet))
    # 获取最大行号
    print(pe.getRowsNumber(sheet))
    # 获取最大列号
    print(pe.getColsNumber(sheet))
    # 获取第一行
    rows = pe.getRow(sheet)
    for i in rows:
        print(i.value)
        # 获取第一行第一列单元格内容
        print(pe.getCellOfValue(sheet, rowNo=1, colsNo=1))
        pe.writeCell(sheet,"我爱祖国", rowNo=10, colsNo=10)
        pe.writeCellCurrentTime(sheet, rowNo=10, colsNo=11)